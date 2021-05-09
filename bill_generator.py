from io import BytesIO
from collections import namedtuple
import configparser
from datetime import datetime
from dateutils import relativedelta
import locale
import logging
import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


CONFIG_NAME = "настройки.txt"
STATEMENT_COLUMNS = ['NUMBER', 'NAME', 'ACCOUNT', 'DEBT', 'DEBT_MONTHS', 'METER_LAST', 'METER_PAID']


def exception(msg, e=None):
    logger = logging.getLogger()
    logger.exception(msg)
    input("Нажмите Enter, чтобы выйти.")
    raise e


class BillGenerator:
    StatementCols = namedtuple(
        'StatementCols',
        STATEMENT_COLUMNS
    )

    def __init__(self, config, statement_columns):
        self.config = config
        self.statement_columns = self.StatementCols(*statement_columns)

    def read_statement(self):

        template_filename = os.path.join(
            self.config['TEMPLATE_FOLDER'],
            self.config['TEMPLATE_FILENAME']
        )
        statement_folder = self.config['STATEMENT_FOLDER']
        first_row = int(self.config['FIRST_ROW'])
        last_row = int(self.config['LAST_ROW'])

        with open(template_filename, 'rb') as file:
            template_bytes = BytesIO(file.read())
        bill_data = []

        statement_files = [file for file in os.listdir(statement_folder) if file.endswith(".xlsx")]
        try:
            statement_filename = statement_files[0]
        except IndexError:
            exception(f"В папке {statement_folder} нет файлов")
        statement_filename_full = os.path.join(statement_folder, statement_filename)

        statement = load_workbook(filename=statement_filename_full, data_only=True).worksheets[0]

        month = statement[self.config['MONTH_CELL']].value
        year = statement[self.config['YEAR_CELL']].value
        date_string = '{month} {year}'.format(month=month, year=year)
        today = datetime.strptime(date_string, '%B %Y')
        for row_index in range(first_row, last_row + 1):
            row = statement[row_index]
            if self.is_valid(row):
                debt = float(self.get_value(row, self.statement_columns.DEBT))
                debt_months = int(self.get_value(row, self.statement_columns.DEBT_MONTHS))
                bill_months = self.get_bill_months(today, debt_months)

                context = {
                    '{%номер%}': self.get_value(row, self.statement_columns.NUMBER),
                    '{%имя%}': self.get_value(row, self.statement_columns.NAME),
                    '{%лицевой_счет%}': self.get_value(row, self.statement_columns.ACCOUNT),
                    '{%месяц%}': bill_months,
                    '{%год%}': year,
                    '{%долг%}': ("%.2f" % debt).replace('.', ','),
                    '{%долг_рубли%}': "%.f" % debt,
                    '{%долг_копейки%}': '0',
                    '{%последнее_показание_1%}': self.get_value(row, self.statement_columns.METER_LAST) or 0,
                    '{%предыдущее_оплаченное_1%}': self.get_value(row, self.statement_columns.METER_PAID) or 0,
                    '{%последнее_показание_2%}': '',
                    '{%предыдущее_оплаченное_2%}': '',
                    '{%последнее_показание_3%}': '',
                    '{%предыдущее_оплаченное_3%}': '',
                    '{%_meters%}': 1,  # should be 1 by default
                    '{%_debt_months%}': debt_months,
                }
                bill_data.append(context)
            elif self.is_second_meter(row):
                try:
                    previous_context = bill_data[-1]
                    meter_index = previous_context['{%_meters%}'] + 1
                    previous_context.update({
                        '{%%последнее_показание_%s%%}' % meter_index:
                            self.get_value(row, self.statement_columns.METER_LAST) or 0,
                        '{%%предыдущее_оплаченное_%s%%}' % meter_index:
                            self.get_value(row, self.statement_columns.METER_PAID) or 0,
                    })
                    previous_context['{%_meters%}'] += 1
                except IndexError:
                    continue

        return template_bytes, bill_data

    @staticmethod
    def get_value(row, column):
        col_index = column_index_from_string(column) - 1
        return row[col_index].value

    @staticmethod
    def get_bill_months(today, months):
        current_month = today.strftime('%B')
        if months <= 1:
            return current_month

        first_month = (today - relativedelta(months=months - 1)).strftime('%B')
        return '%s - %s' % (first_month, current_month)

    def fill_template(self, template_bytes, context):
        if int(context['{%_debt_months%}']) >= int(self.config['DEBT_MONTHS']):
            template_wb = load_workbook(template_bytes)
            sheet = template_wb.worksheets[0]

            output_filename = self.config['OUTPUT_FILENAME_FORMAT']
            output_folder = self.config['OUTPUT_FOLDER']
            for row in sheet:
                for cell in row:
                    for key, value in context.items():
                        if key in str(cell.value):
                            cell.value = cell.value.replace(key, str(value))
                        output_filename = output_filename.replace(key, str(value))
            output_filename_full = os.path.join(output_folder, output_filename)
            template_wb.save(filename=output_filename_full)
            logging.info(f"Сделано: {output_filename}")

    def is_valid(self, row):
        required_cols = [
            self.statement_columns.NUMBER,
            self.statement_columns.NAME,
            self.statement_columns.DEBT_MONTHS,
        ]
        if all([self.get_value(row, col) is not None for col in required_cols]):
            return True
        return False

    def is_second_meter(self, row):
        required_cols = [
            self.statement_columns.METER_LAST,
            self.statement_columns.METER_PAID,
        ]
        required_cols_false = [
            self.statement_columns.NAME,
        ]
        if all([self.get_value(row, col) is not None for col in required_cols]) \
                and not any(self.get_value(row, col) for col in required_cols_false):
            return True
        return False


class App:
    def __init__(self):
        self.exit_flag = False
        locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
        if os.path.exists(CONFIG_NAME):
            self.config, self.statement_columns = self.read_config()
        else:
            self.config = self.generate_default_config()
            self.exit_flag = True
        folders_list = [
            'STATEMENT_FOLDER',
            'OUTPUT_FOLDER',
            'TEMPLATE_FOLDER',
            'LOG_FOLDER',
        ]
        for folder in folders_list:
            if not os.path.exists(self.config[folder]):
                os.mkdir(self.config[folder])
                self.exit_flag = True
        self.init_logger()

    @staticmethod
    def generate_default_config():
        config = configparser.RawConfigParser(allow_no_value=True)
        config.optionxform = str
        config.add_section('SETTINGS')
        config.set('SETTINGS', '# Папка с ведомостью')
        config.set('SETTINGS', 'STATEMENT_FOLDER', 'Ведомость')
        config.set('SETTINGS', '')
        config.set('SETTINGS', '# Папка с логами работы программы')
        config.set('SETTINGS', 'LOG_FOLDER', 'log')
        config.set('SETTINGS', '')
        config.set('SETTINGS', '# Папка с шаблоном квитанции и имя самого файла-шаблона')
        config.set('SETTINGS', 'TEMPLATE_FOLDER', 'Шаблон')
        config.set('SETTINGS', 'TEMPLATE_FILENAME', 'квитанция.xlsx')
        config.set('SETTINGS', ' ')
        config.set('SETTINGS', '# Папка, куда будут сложены все квитанции')
        config.set('SETTINGS', 'OUTPUT_FOLDER', 'Квитанции')
        config.set('SETTINGS', '  ')
        config.set('SETTINGS', '# Формат имени выходного файла')
        config.set('SETTINGS', 'OUTPUT_FILENAME_FORMAT', '{%номер%}_{%месяц%}_{%имя%}.xlsx')
        config.set('SETTINGS', '   ')
        config.set('SETTINGS', '# Первый и последний ряды в ведомости')
        config.set('SETTINGS', 'FIRST_ROW', '9')
        config.set('SETTINGS', 'LAST_ROW', '170')
        config.set('SETTINGS', '    ')
        config.set('SETTINGS', '# Количество месяцев просроченных платежей')
        config.set('SETTINGS', 'DEBT_MONTHS', '3')
        config.set('SETTINGS', '# Ячейка с месяцем')
        config.set('SETTINGS', 'MONTH_CELL', 'I2')
        config.set('SETTINGS', '# Ячейка с годом')
        config.set('SETTINGS', 'YEAR_CELL', 'J2')

        config.add_section('COLUMNS')
        config.set('COLUMNS', '# Столбцы')
        config.set('COLUMNS', '# Номер по порядку')
        config.set('COLUMNS', 'NUMBER', 'A')
        config.set('COLUMNS', '# Имя')
        config.set('COLUMNS', 'NAME', 'C')
        config.set('COLUMNS', '# Номер лицевого счета')
        config.set('COLUMNS', 'ACCOUNT', 'D')
        config.set('COLUMNS', '# Долг')
        config.set('COLUMNS', 'DEBT', 'I')
        config.set('COLUMNS', '# Месяцев задолжность')
        config.set('COLUMNS', 'DEBT_MONTHS', 'J')
        config.set('COLUMNS', '# Последнее показание счетчика')
        config.set('COLUMNS', 'METER_LAST', 'K')
        config.set('COLUMNS', '# Предыдущее оплаченное')
        config.set('COLUMNS', 'METER_PAID', 'L')

        with open(CONFIG_NAME, 'w', encoding='utf-8') as file:
            config.write(file)

        return config['SETTINGS']

    @staticmethod
    def read_config():
        config = configparser.RawConfigParser()
        config.read(CONFIG_NAME, encoding="utf-8")

        cols_conf = config['COLUMNS']

        statement_columns = [cols_conf[col] for col in STATEMENT_COLUMNS]

        return config['SETTINGS'], statement_columns

    def run(self):
        try:
            bill_generator = BillGenerator(self.config, self.statement_columns)
            template_bytes, bill_data = bill_generator.read_statement()
            for context in bill_data:
                bill_generator.fill_template(template_bytes, context)
        except PermissionError:
            exception("Произошла ошибка. Закройте все файлы Excel перед запуском и попробуйте еще раз.")
        except Exception as e:
            exception("Произошла непредвиденная ошибка. Лучше показать это Артёму.", e=e)

    def init_logger(self):
        now = datetime.now()
        logging_filename = os.path.join(self.config['LOG_FOLDER'],
                                        f'log_{now.strftime("%Y-%m-%d_%H-%M-%S")}.txt')

        file_handler = logging.FileHandler(logging_filename, "w", encoding="UTF-8")
        stream_handler = logging.StreamHandler()

        root_logger = logging.getLogger()
        root_logger.addHandler(file_handler)
        root_logger.addHandler(stream_handler)
        root_logger.setLevel(logging.DEBUG)


if __name__ == '__main__':
    app = App()
    if not app.exit_flag:
        app.run()
        logging.info("Успех! Все получилось. Проверьте файлы.")
        input("Нажмите Enter, чтобы выйти.")
    else:
        exception("Приложение инициализировано. "
                  "Вложите файлы в соответствующие папки и запустите еще раз.")
