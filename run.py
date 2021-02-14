import configparser
import os
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


CONFIG_NAME = "настройки.txt"


class BillGenerator:
    class StatementCols:
        NUMBER = 'A'
        NAME = 'C'
        ACCOUNT = 'D'
        DEBT = 'I'

    @staticmethod
    def get_col_index(letter):
        return column_index_from_string(letter) - 1

    def fill_template(self, context):
        sheet = self.template_wb.worksheets[0]

        output_filename = OUTPUT_FILENAME_FORMAT
        for row in sheet:
            for cell in row:
                for key, value in context.items():
                    if key in str(cell.value):
                        cell.value = cell.value.replace(key, str(value))
                    output_filename = output_filename.replace(key, str(value))
        output_filename_full = os.path.join(OUTPUT_FOLDER, output_filename)
        self.template_wb.save(filename=output_filename_full)

    @classmethod
    def is_valid(cls, row):
        if all([bool(row[cls.get_col_index(col)].value) for col in [
                    cls.StatementCols.NUMBER,
                    cls.StatementCols.NAME,
                ]]):
            return True
        return False

    def __init__(self, template_wb):
        self.template_wb = template_wb
        self.bill_data = []

        statement_files = [file for file in os.listdir(STATEMENT_FOLDER) if file.endswith(".xlsx")]
        try:
            statement_filename = statement_files[0]
        except IndexError:
            print(f"В папке {STATEMENT_FOLDER} нет файлов")
            exit()
        statement_filename_full = os.path.join(STATEMENT_FOLDER, statement_filename)

        statement = load_workbook(filename=statement_filename_full, data_only=True).worksheets[0]
        for row_index in range(FIRST_ROW, LAST_ROW + 1):
            row = statement[row_index]
            if self.is_valid(row):
                context = {
                    '{%номер%}': row[self.get_col_index(self.StatementCols.NUMBER)].value,
                    '{%имя%}': row[self.get_col_index(self.StatementCols.NAME)].value,
                    '{%лицевой_счет%}': row[self.get_col_index(self.StatementCols.ACCOUNT)].value,
                    '{%месяц%}': 'Февраль',
                    '{%год%}': '2021',
                    '{%долг%}': row[self.get_col_index(self.StatementCols.DEBT)].value,
                    '{%долг_рубли%}': '4043',
                    '{%долг_копейки%}': '00',
                }
                self.bill_data.append(context)


if __name__ == '__main__':
    config = configparser.RawConfigParser()
    config.read(CONFIG_NAME, encoding="utf-8")
    structure_conf = config['STRUCTURE']

    STATEMENT_FOLDER = structure_conf['STATEMENT_FOLDER']

    TEMPLATE_FOLDER = structure_conf['TEMPLATE_FOLDER']
    TEMPLATE_FILENAME = structure_conf['TEMPLATE_FILENAME']

    OUTPUT_FOLDER = structure_conf['OUTPUT_FOLDER']
    OUTPUT_FILENAME_FORMAT = structure_conf['OUTPUT_FILENAME_FORMAT']

    FIRST_ROW = int(structure_conf['FIRST_ROW'])
    LAST_ROW = int(structure_conf['LAST_ROW'])

    cols_conf = config['COLUMNS']

    BillGenerator.StatementCols.NAME = cols_conf['NAME']
    BillGenerator.StatementCols.NUMBER = cols_conf['NUMBER']
    BillGenerator.StatementCols.DEBT = cols_conf['DEBT']
    BillGenerator.StatementCols.ACCOUNT = cols_conf['ACCOUNT']

    try:
        template_filename = os.path.join(TEMPLATE_FOLDER, TEMPLATE_FILENAME)
        template_wb = load_workbook(filename=template_filename)

        bill_generator = BillGenerator(template_wb)

        for context in bill_generator.bill_data:
            bill_generator.fill_template(context)
    except PermissionError:
        print("Произошла ошибка. Закройте все файлы Excel перед запуском.")
    except:
        print("Произошла непредвиденная ошибка.")
        raise
